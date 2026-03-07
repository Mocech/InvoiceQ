"""
Export Service — CSV, JSON, Excel, and PDF exports of invoices.

Site brand colors (from styles.css):
  --primary:      #1E3A5F   (navy — headers, banners)
  --sidebar-dark: #0F1B2D   (deep navy — title bars)
  --success:      #27AE60   (green — approved status)
  --warning:      #F39C12   (amber — pending/flagged)
  --danger:       #E74C3C   (red — rejected)
  --bg-light:     #F4F6F9   (page background — alternating rows)
  --white:        #FFFFFF
  --text-dark:    #2D3748   (body text)
  --text-light:   #718096   (labels/subtext)
  --border-light: #E2E8F0   (borders)
"""

import csv
import json
import io
from datetime import date, datetime

from django.http import HttpResponse


class ExportService:

    # ── CSV ───────────────────────────────────────────────────────────────────

    def to_csv(self, invoice) -> HttpResponse:
        output = io.StringIO()
        writer = csv.writer(output)

        writer.writerow(['InvoiceIQ Export'])
        writer.writerow([])
        writer.writerow(['Field', 'Value'])
        writer.writerow(['Invoice Number',  invoice.invoice_number])
        writer.writerow(['Vendor Name',     invoice.vendor_name])
        writer.writerow(['Vendor Address',  invoice.vendor_address])
        writer.writerow(['Vendor Email',    invoice.vendor_email])
        writer.writerow(['Vendor Phone',    invoice.vendor_phone])
        writer.writerow(['PO Number',       invoice.po_number])
        writer.writerow(['Invoice Date',    invoice.invoice_date])
        writer.writerow(['Due Date',        invoice.due_date])
        writer.writerow(['Payment Terms',   invoice.payment_terms])
        writer.writerow(['Currency',        invoice.currency])
        writer.writerow(['Subtotal',        invoice.subtotal_amount])
        writer.writerow(['Tax',             invoice.tax_amount])
        writer.writerow(['Total Amount',    invoice.total_amount])
        writer.writerow(['Approval Status', (invoice.status or '').replace('_', ' ').title()])
        writer.writerow([])

        writer.writerow(['Line Items'])
        writer.writerow(['Description', 'Quantity', 'Unit Price', 'Total'])
        for item in invoice.line_items.all():
            writer.writerow([item.description, item.quantity, item.unit_price, item.total])

        writer.writerow([])
        writer.writerow(['Audit Trail'])
        writer.writerow(['Action', 'Performed By', 'Date & Time'])
        for log in invoice.audit_logs.select_related('user').order_by('timestamp'):
            writer.writerow([
                log.action.replace('_', ' ').title(),
                log.user.full_name if log.user else 'System',
                log.timestamp.strftime('%d %b %Y %H:%M'),
            ])

        content  = output.getvalue()
        response = HttpResponse(content, content_type='text/csv')
        filename = f'invoice_{invoice.invoice_number or invoice.id}.csv'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    # ── JSON ──────────────────────────────────────────────────────────────────

    def to_json(self, invoice) -> HttpResponse:
        data = {
            'export_source':  'InvoiceIQ',
            'export_version': '1.0',
            'invoice': {
                'id':             invoice.id,
                'invoice_number': invoice.invoice_number,
                'status':         invoice.status,
                'vendor': {
                    'name':    invoice.vendor_name,
                    'address': invoice.vendor_address,
                    'email':   invoice.vendor_email,
                    'phone':   invoice.vendor_phone,
                },
                'po_number':     invoice.po_number,
                'payment_terms': invoice.payment_terms,
                'dates': {
                    'invoice_date': str(invoice.invoice_date) if invoice.invoice_date else None,
                    'due_date':     str(invoice.due_date)     if invoice.due_date     else None,
                },
                'amounts': {
                    'currency': invoice.currency,
                    'subtotal': str(invoice.subtotal_amount or 0),
                    'tax':      str(invoice.tax_amount      or 0),
                    'total':    str(invoice.total_amount    or 0),
                },
                'line_items': [
                    {
                        'description': item.description,
                        'quantity':    str(item.quantity),
                        'unit_price':  str(item.unit_price),
                        'total':       str(item.total),
                    }
                    for item in invoice.line_items.all()
                ],
                'audit_trail': [
                    {
                        'action':       log.action,
                        'performed_by': log.user.full_name if log.user else 'System',
                        'timestamp':    log.timestamp.isoformat(),
                    }
                    for log in invoice.audit_logs.select_related('user').order_by('timestamp')
                ],
            }
        }
        response = HttpResponse(json.dumps(data, indent=2), content_type='application/json')
        filename = f'invoice_{invoice.invoice_number or invoice.id}.json'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    # ── EXCEL ─────────────────────────────────────────────────────────────────

    def to_excel(self, invoice) -> HttpResponse:
        """
        Sheet 1 — Invoice Details: vendor, dates, amounts, approval, audit trail
        Sheet 2 — Line Items: flat accounting-ready table
        Colors match site brand palette from styles.css
        """
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        wb = openpyxl.Workbook()

        # ── Brand colors (exact from styles.css) ─────────────────────────
        PRIMARY     = '1E3A5F'   # --primary (navy)
        SIDEBAR     = '0F1B2D'   # --sidebar-dark (deep navy)
        SUCCESS     = '27AE60'   # --success (green)
        WARNING     = 'F39C12'   # --warning (amber)
        DANGER      = 'E74C3C'   # --danger (red)
        BG_LIGHT    = 'F4F6F9'   # --bg-light (alternating rows)
        TEXT_DARK   = '2D3748'   # --text-dark
        TEXT_LIGHT  = '718096'   # --text-light (labels)
        BORDER      = 'E2E8F0'   # --border-light
        WHITE       = 'FFFFFF'
        # Accent blue for column headers (slightly lighter than primary)
        ACCENT      = '2C4F7C'

        def _bdr():
            s = Side(style='thin', color=BORDER)
            return Border(left=s, right=s, top=s, bottom=s)

        def _c(ws, row, col, value, bold=False, size=10,
               fg=TEXT_DARK, bg=None, align='left', wrap=False, num_fmt=None):
            c = ws.cell(row=row, column=col, value=value)
            c.font      = Font(name='Calibri', bold=bold, size=size, color=fg)
            c.alignment = Alignment(horizontal=align, vertical='center', wrap_text=wrap)
            if bg:
                c.fill = PatternFill('solid', fgColor=bg)
            if num_fmt:
                c.number_format = num_fmt
            c.border = _bdr()
            return c

        def _section(ws, row, text, end_col='D'):
            ws.merge_cells(f'A{row}:{end_col}{row}')
            c = ws.cell(row=row, column=1, value=text)
            c.font      = Font(name='Calibri', bold=True, size=9, color=WHITE)
            c.fill      = PatternFill('solid', fgColor=ACCENT)
            c.alignment = Alignment(horizontal='left', vertical='center', indent=1)
            ws.row_dimensions[row].height = 20
            return row + 1

        def _kv_row(ws, row, label, value, i, end_col='D', num_fmt=None):
            """Alternating key-value row spanning columns."""
            bg = BG_LIGHT if i % 2 == 0 else WHITE
            _c(ws, row, 1, label, bold=True, bg=bg, fg=TEXT_LIGHT, size=9)
            _c(ws, row, 2, value or '—', bg=bg, fg=TEXT_DARK,
               num_fmt=num_fmt if num_fmt else None)
            ws.merge_cells(f'B{row}:{end_col}{row}')
            ws.row_dimensions[row].height = 18

        # ── Status helpers ────────────────────────────────────────────────
        status_raw   = (invoice.status or '').lower()
        status_label = status_raw.replace('_', ' ').title()
        status_color = (SUCCESS if status_raw == 'approved' else
                        DANGER  if status_raw == 'rejected' else WARNING)
        currency     = invoice.currency or ''

        # ─────────────────────────────────────────────────────────────────
        # SHEET 1 — Invoice Details
        # ─────────────────────────────────────────────────────────────────
        ws1 = wb.active
        ws1.title = 'Invoice Details'
        ws1.sheet_view.showGridLines = False
        ws1.column_dimensions['A'].width = 22
        ws1.column_dimensions['B'].width = 32
        ws1.column_dimensions['C'].width = 18
        ws1.column_dimensions['D'].width = 18

        row = 1

        # Title banner — sidebar-dark background
        ws1.merge_cells(f'A{row}:D{row}')
        c = ws1.cell(row=row, column=1, value='INVOICE RECORD')
        c.font      = Font(name='Calibri', bold=True, size=16, color=WHITE)
        c.fill      = PatternFill('solid', fgColor=SIDEBAR)
        c.alignment = Alignment(horizontal='left', vertical='center', indent=2)
        ws1.row_dimensions[row].height = 40
        row += 1

        # Export date + status bar
        ws1.merge_cells(f'A{row}:B{row}')
        c = ws1.cell(row=row, column=1,
                     value=f'Exported: {datetime.now().strftime("%d %b %Y  %H:%M")}')
        c.font      = Font(name='Calibri', size=9, color=TEXT_LIGHT)
        c.fill      = PatternFill('solid', fgColor=BG_LIGHT)
        c.alignment = Alignment(horizontal='left', vertical='center', indent=2)
        c.border    = _bdr()

        ws1.merge_cells(f'C{row}:D{row}')
        c2 = ws1.cell(row=row, column=3, value=f'● {status_label.upper()}')
        c2.font      = Font(name='Calibri', bold=True, size=10, color=WHITE)
        c2.fill      = PatternFill('solid', fgColor=status_color)
        c2.alignment = Alignment(horizontal='center', vertical='center')
        c2.border    = _bdr()
        ws1.row_dimensions[row].height = 22
        row += 2

        # ── Vendor ───────────────────────────────────────────────────────
        row = _section(ws1, row, 'VENDOR INFORMATION')
        for i, (lbl, val) in enumerate([
            ('Vendor Name',  invoice.vendor_name),
            ('Address',      invoice.vendor_address),
            ('Email',        invoice.vendor_email),
            ('Phone',        invoice.vendor_phone),
        ]):
            _kv_row(ws1, row, lbl, val, i)
            row += 1
        row += 1

        # ── Invoice Info ─────────────────────────────────────────────────
        row = _section(ws1, row, 'INVOICE INFORMATION')
        for i, (lbl, val) in enumerate([
            ('Invoice Number', invoice.invoice_number),
            ('PO Number',      invoice.po_number or '—'),
            ('Payment Terms',  invoice.payment_terms),
            ('Invoice Date',   str(invoice.invoice_date) if invoice.invoice_date else '—'),
            ('Due Date',       str(invoice.due_date)     if invoice.due_date     else '—'),
            ('Currency',       invoice.currency),
        ]):
            _kv_row(ws1, row, lbl, val, i)
            row += 1
        row += 1

        # ── Amounts ──────────────────────────────────────────────────────
        row = _section(ws1, row, 'AMOUNTS')
        amounts = [
            ('Subtotal',     float(invoice.subtotal_amount or 0)),
            ('Tax',          float(invoice.tax_amount      or 0)),
            ('Total Amount', float(invoice.total_amount    or 0)),
        ]
        for i, (lbl, amt) in enumerate(amounts):
            is_total = lbl == 'Total Amount'
            bg = PRIMARY if is_total else (BG_LIGHT if i % 2 == 0 else WHITE)
            fg = WHITE   if is_total else TEXT_DARK
            _c(ws1, row, 1, lbl, bold=is_total, bg=bg, fg=fg, size=10 if is_total else 9)
            _c(ws1, row, 2, amt, bold=is_total, bg=bg, fg=fg,
               align='right', num_fmt=f'"{currency} "#,##0.00')
            ws1.merge_cells(f'B{row}:D{row}')
            ws1.row_dimensions[row].height = 24 if is_total else 18
            row += 1
        row += 1

        # ── Approval ─────────────────────────────────────────────────────
        row = _section(ws1, row, 'APPROVAL')
        approved_log = (invoice.audit_logs
                        .filter(action='approved')
                        .select_related('user').first())
        for i, (lbl, val) in enumerate([
            ('Approval Status', status_label),
            ('Approved By',     approved_log.user.full_name
                                if approved_log and approved_log.user else '—'),
            ('Approved On',     approved_log.timestamp.strftime('%d %b %Y %H:%M')
                                if approved_log else '—'),
        ]):
            _kv_row(ws1, row, lbl, val, i)
            row += 1
        row += 1

        # ── Audit Trail ──────────────────────────────────────────────────
        audit_logs = list(invoice.audit_logs.select_related('user').order_by('timestamp'))
        if audit_logs:
            row = _section(ws1, row, 'AUDIT TRAIL')
            # Column headers
            for col, txt in enumerate(['Action', 'Performed By', 'Date & Time'], 1):
                _c(ws1, row, col, txt, bold=True, bg=BG_LIGHT,
                   fg=TEXT_LIGHT, size=9, align='left')
            ws1.merge_cells(f'C{row}:D{row}')
            ws1.row_dimensions[row].height = 18
            row += 1
            for i, log in enumerate(audit_logs):
                bg = BG_LIGHT if i % 2 == 0 else WHITE
                _c(ws1, row, 1, log.action.replace('_', ' ').title(), bg=bg, size=9)
                _c(ws1, row, 2, log.user.full_name if log.user else 'System', bg=bg, size=9)
                _c(ws1, row, 3, log.timestamp.strftime('%d %b %Y %H:%M'), bg=bg, size=9)
                ws1.merge_cells(f'C{row}:D{row}')
                ws1.row_dimensions[row].height = 18
                row += 1

        # ── Generated by footer ──────────────────────────────────────────
        row += 1
        ws1.merge_cells(f'A{row}:D{row}')
        c = ws1.cell(row=row, column=1,
                     value=f'Generated by InvoiceIQ  |  Invoice ID: {invoice.id}  |  {datetime.now().strftime("%d %b %Y %H:%M")}')
        c.font      = Font(name='Calibri', size=8, color=TEXT_LIGHT, italic=True)
        c.alignment = Alignment(horizontal='left', vertical='center', indent=1)
        c.fill      = PatternFill('solid', fgColor=BG_LIGHT)
        c.border    = _bdr()
        ws1.row_dimensions[row].height = 16

        # ─────────────────────────────────────────────────────────────────
        # SHEET 2 — Line Items (flat, accounting-ready)
        # ─────────────────────────────────────────────────────────────────
        ws2 = wb.create_sheet(title='Line Items')
        ws2.sheet_view.showGridLines = False
        ws2.column_dimensions['A'].width = 16
        ws2.column_dimensions['B'].width = 20
        ws2.column_dimensions['C'].width = 13
        ws2.column_dimensions['D'].width = 13
        ws2.column_dimensions['E'].width = 38
        ws2.column_dimensions['F'].width = 8
        ws2.column_dimensions['G'].width = 14
        ws2.column_dimensions['H'].width = 14
        ws2.column_dimensions['I'].width = 10
        ws2.column_dimensions['J'].width = 14

        r = 1

        # Title
        ws2.merge_cells(f'A{r}:J{r}')
        c = ws2.cell(row=r, column=1,
                     value=f'{invoice.invoice_number or "Invoice"}  |  '
                           f'{invoice.vendor_name or ""}  |  '
                           f'{status_label.upper()}')
        c.font      = Font(name='Calibri', bold=True, size=12, color=WHITE)
        c.fill      = PatternFill('solid', fgColor=SIDEBAR)
        c.alignment = Alignment(horizontal='left', vertical='center', indent=2)
        ws2.row_dimensions[r].height = 32
        r += 1

        # Column headers
        for col, txt in enumerate([
            'Invoice No', 'Vendor', 'Invoice Date', 'Due Date',
            'Description', 'Qty', 'Unit Price', 'Line Total', 'Currency', 'Status'
        ], 1):
            _c(ws2, r, col, txt, bold=True, bg=PRIMARY, fg=WHITE, align='center', size=9)
        ws2.row_dimensions[r].height = 20
        r += 1

        # Data rows
        inv_no   = invoice.invoice_number or str(invoice.id)
        inv_date = str(invoice.invoice_date) if invoice.invoice_date else ''
        inv_due  = str(invoice.due_date)     if invoice.due_date     else ''

        line_items = list(invoice.line_items.all().order_by('sort_order'))
        for i, item in enumerate(line_items):
            bg = BG_LIGHT if i % 2 == 0 else WHITE
            _c(ws2, r, 1,  inv_no,                       bg=bg, size=9)
            _c(ws2, r, 2,  invoice.vendor_name or '',     bg=bg, size=9)
            _c(ws2, r, 3,  inv_date,                      bg=bg, size=9, align='center')
            _c(ws2, r, 4,  inv_due,                       bg=bg, size=9, align='center')
            _c(ws2, r, 5,  item.description or '',        bg=bg, size=9, wrap=True)
            _c(ws2, r, 6,  float(item.quantity  or 0),    bg=bg, size=9, align='center')
            _c(ws2, r, 7,  float(item.unit_price or 0),   bg=bg, size=9, align='right',
               num_fmt='#,##0.00')
            _c(ws2, r, 8,  float(item.total     or 0),    bg=bg, size=9, align='right',
               num_fmt='#,##0.00')
            _c(ws2, r, 9,  currency,                      bg=bg, size=9, align='center')
            _c(ws2, r, 10, status_label,                  bg=bg, size=9, align='center')
            ws2.row_dimensions[r].height = 18
            r += 1

        # Totals
        r += 1
        for lbl, amt in [
            ('Subtotal', float(invoice.subtotal_amount or 0)),
            ('Tax',      float(invoice.tax_amount      or 0)),
            ('TOTAL',    float(invoice.total_amount    or 0)),
        ]:
            is_total = lbl == 'TOTAL'
            bg = PRIMARY if is_total else BG_LIGHT
            fg = WHITE   if is_total else TEXT_DARK
            ws2.merge_cells(f'A{r}:G{r}')
            _c(ws2, r, 1, lbl,      bold=is_total, bg=bg, fg=fg,
               align='right', size=10 if is_total else 9)
            _c(ws2, r, 8, amt,      bold=is_total, bg=bg, fg=fg,
               align='right', num_fmt='#,##0.00')
            _c(ws2, r, 9, currency, bg=bg, fg=fg, align='center', size=9)
            _c(ws2, r, 10, '',      bg=bg)
            ws2.row_dimensions[r].height = 24 if is_total else 18
            r += 1

        # Generated by footer
        r += 1
        ws2.merge_cells(f'A{r}:J{r}')
        c = ws2.cell(row=r, column=1,
                     value=f'Generated by InvoiceIQ  |  Invoice ID: {invoice.id}  |  {datetime.now().strftime("%d %b %Y %H:%M")}')
        c.font      = Font(name='Calibri', size=8, color=TEXT_LIGHT, italic=True)
        c.alignment = Alignment(horizontal='left', vertical='center', indent=1)
        c.fill      = PatternFill('solid', fgColor=BG_LIGHT)
        c.border    = _bdr()
        ws2.row_dimensions[r].height = 16

        # ── Save ─────────────────────────────────────────────────────────
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        filename = f'invoice_{invoice.invoice_number or invoice.id}.xlsx'
        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    # ── PDF ───────────────────────────────────────────────────────────────────

    def to_pdf(self, invoice) -> HttpResponse:
        """
        Professional printable PDF using exact site brand colors.
        Layer 1: Header (InvoiceIQ branding + invoice number + status)
        Layer 2: Vendor info + invoice details (two-column)
        Layer 3: Line items table
        Layer 4: Approval block
        Layer 5: Audit trail (with Generated By)
        Footer:  Generated by InvoiceIQ
        """
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.lib.units import mm
        from reportlab.platypus import (
            SimpleDocTemplate, Paragraph, Spacer,
            Table, TableStyle, HRFlowable, KeepTogether
        )
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_RIGHT

        # ── Brand colors (exact from styles.css) ─────────────────────────
        PRIMARY  = colors.HexColor('#1E3A5F')   # --primary
        SIDEBAR  = colors.HexColor('#0F1B2D')   # --sidebar-dark
        SUCCESS  = colors.HexColor('#27AE60')   # --success
        WARNING  = colors.HexColor('#F39C12')   # --warning
        DANGER   = colors.HexColor('#E74C3C')   # --danger
        BG_LIGHT = colors.HexColor('#F4F6F9')   # --bg-light
        TXT_DARK = colors.HexColor('#2D3748')   # --text-dark
        TXT_LITE = colors.HexColor('#718096')   # --text-light
        BORDER   = colors.HexColor('#E2E8F0')   # --border-light
        WHITE    = colors.white
        ACCENT   = colors.HexColor('#2C4F7C')   # slightly lighter primary

        buffer    = io.BytesIO()
        margin    = 18 * mm
        page_w, _ = A4
        usable_w  = page_w - 2 * margin

        doc = SimpleDocTemplate(
            buffer, pagesize=A4,
            leftMargin=margin, rightMargin=margin,
            topMargin=margin, bottomMargin=margin,
        )

        base = getSampleStyleSheet()['Normal']

        def _s(name, **kw):
            return ParagraphStyle(name, parent=base, **kw)

        # Style definitions using brand colors
        lbl_s   = _s('lbl',  fontSize=7,  textColor=TXT_LITE, fontName='Helvetica-Bold',
                      leading=10, spaceAfter=2)
        val_s   = _s('val',  fontSize=9,  textColor=TXT_DARK, fontName='Helvetica',
                      leading=13)
        bld_s   = _s('bld',  fontSize=9,  textColor=TXT_DARK, fontName='Helvetica-Bold',
                      leading=13)
        sm_s    = _s('sm',   fontSize=8,  textColor=TXT_LITE, fontName='Helvetica',
                      leading=11)
        h2_s    = _s('h2',   fontSize=10, textColor=PRIMARY,  fontName='Helvetica-Bold',
                      leading=14, spaceBefore=8, spaceAfter=4)
        right_s = _s('r',    fontSize=9,  textColor=TXT_DARK, fontName='Helvetica',
                      leading=13, alignment=TA_RIGHT)
        right_b = _s('rb',   fontSize=9,  textColor=TXT_DARK, fontName='Helvetica-Bold',
                      leading=13, alignment=TA_RIGHT)
        ctr_s   = _s('ctr',  fontSize=9,  textColor=TXT_DARK, fontName='Helvetica',
                      leading=13, alignment=TA_CENTER)

        elements  = []
        currency  = invoice.currency or ''
        status_raw = (invoice.status or '').lower()
        status_lbl = status_raw.replace('_', ' ').title()
        status_clr = (SUCCESS if status_raw == 'approved' else
                      DANGER  if status_raw == 'rejected' else WARNING)
        status_hex = ('#27AE60' if status_raw == 'approved' else
                      '#E74C3C' if status_raw == 'rejected' else '#F39C12')
        inv_no     = invoice.invoice_number or f'ID-{invoice.id}'

        # ── HEADER ───────────────────────────────────────────────────────
        # Deep navy (sidebar-dark) background — matches sidebar branding
        header = Table([[
            Paragraph(
                '<font color="white" size="18"><b>InvoiceIQ</b></font><br/>'
                '<font color="#718096" size="8">Invoice Processing Platform</font>',
                _s('hl', fontSize=9, textColor=WHITE, leading=18)
            ),
            Paragraph(
                f'<font color="#718096" size="7">INVOICE NUMBER</font><br/>'
                f'<font color="white" size="15"><b>{inv_no}</b></font><br/>'
                f'<font color="{status_hex}" size="9"><b>● {status_lbl.upper()}</b></font>',
                _s('hr', fontSize=9, textColor=WHITE, leading=16, alignment=TA_RIGHT)
            ),
        ]], colWidths=[usable_w * 0.55, usable_w * 0.45])
        header.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), SIDEBAR),
            ('PADDING',    (0, 0), (-1, -1), 16),
            ('VALIGN',     (0, 0), (-1, -1), 'MIDDLE'),
            ('ROUNDEDCORNERS', [4]),
        ]))
        elements.append(header)
        elements.append(Spacer(1, 7 * mm))

        # ── VENDOR + INVOICE INFO (two-column card) ───────────────────────
        def _lv(label, value):
            return [Paragraph(label, lbl_s), Paragraph(str(value or '—'), val_s)]

        left_rows  = [
            _lv('VENDOR',        invoice.vendor_name),
            _lv('ADDRESS',       invoice.vendor_address),
            _lv('EMAIL',         invoice.vendor_email),
            _lv('PHONE',         invoice.vendor_phone),
        ]
        right_rows = [
            _lv('INVOICE DATE',  str(invoice.invoice_date) if invoice.invoice_date else '—'),
            _lv('DUE DATE',      str(invoice.due_date)     if invoice.due_date     else '—'),
            _lv('PAYMENT TERMS', invoice.payment_terms),
            _lv('PO NUMBER',     invoice.po_number or '—'),
        ]

        def _info_col(rows):
            t = Table(rows, colWidths=[usable_w * 0.46])
            t.setStyle(TableStyle([
                ('PADDING', (0, 0), (-1, -1), 5),
                ('VALIGN',  (0, 0), (-1, -1), 'TOP'),
            ]))
            return t

        two_col = Table(
            [[_info_col(left_rows), _info_col(right_rows)]],
            colWidths=[usable_w * 0.5, usable_w * 0.5]
        )
        two_col.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), BG_LIGHT),
            ('BOX',        (0, 0), (-1, -1), 0.5, BORDER),
            ('LINEAFTER',  (0, 0), (0, -1),  0.5, BORDER),
            ('PADDING',    (0, 0), (-1, -1), 12),
            ('VALIGN',     (0, 0), (-1, -1), 'TOP'),
        ]))
        elements.append(two_col)
        elements.append(Spacer(1, 7 * mm))

        # ── LINE ITEMS ────────────────────────────────────────────────────
        line_items = list(invoice.line_items.all().order_by('sort_order'))
        if line_items:
            elements.append(Paragraph('LINE ITEMS', h2_s))
            rows = [[
                Paragraph('<b>Description</b>', bld_s),
                Paragraph('<b>Qty</b>',
                          _s('qh', fontSize=9, fontName='Helvetica-Bold', alignment=TA_CENTER)),
                Paragraph('<b>Unit Price</b>',
                          _s('uh', fontSize=9, fontName='Helvetica-Bold', alignment=TA_RIGHT)),
                Paragraph('<b>Total</b>',
                          _s('th', fontSize=9, fontName='Helvetica-Bold', alignment=TA_RIGHT)),
            ]]

            subtotal = 0
            for item in line_items:
                total     = float(item.total or 0)
                subtotal += total
                rows.append([
                    Paragraph(item.description or '', val_s),
                    Paragraph(str(item.quantity or ''), ctr_s),
                    Paragraph(f'{currency} {float(item.unit_price or 0):,.2f}', right_s),
                    Paragraph(f'{currency} {total:,.2f}', right_s),
                ])

            tax        = float(invoice.tax_amount   or 0)
            grand_total = float(invoice.total_amount or 0)
            empty      = Paragraph('', val_s)

            rows += [
                [empty, empty, Paragraph('Subtotal', right_b),
                 Paragraph(f'{currency} {subtotal:,.2f}', right_s)],
                [empty, empty, Paragraph('Tax', right_b),
                 Paragraph(f'{currency} {tax:,.2f}', right_s)],
                [Paragraph('', _s('we', textColor=WHITE)),
                 Paragraph('', _s('we2', textColor=WHITE)),
                 Paragraph('<b>TOTAL DUE</b>',
                           _s('tdl', fontSize=10, fontName='Helvetica-Bold',
                              alignment=TA_RIGHT, textColor=WHITE)),
                 Paragraph(f'<b>{currency} {grand_total:,.2f}</b>',
                           _s('tda', fontSize=10, fontName='Helvetica-Bold',
                              alignment=TA_RIGHT, textColor=WHITE))],
            ]

            n   = len(rows)
            tbl = Table(rows, colWidths=[usable_w*0.46, usable_w*0.1,
                                         usable_w*0.22, usable_w*0.22])
            tbl.setStyle(TableStyle([
                # Header row — primary navy
                ('BACKGROUND',    (0, 0),   (-1, 0),     PRIMARY),
                ('TEXTCOLOR',     (0, 0),   (-1, 0),     WHITE),
                # Data rows — bg-light alternating
                ('ROWBACKGROUNDS',(0, 1),   (-1, n-4),   [BG_LIGHT, WHITE]),
                # Subtotal / Tax rows — white
                ('BACKGROUND',    (0, n-3), (-1, n-2),   WHITE),
                # Total row — sidebar-dark
                ('BACKGROUND',    (0, n-1), (-1, n-1),   SIDEBAR),
                ('GRID',          (0, 0),   (-1, n-4),   0.5, BORDER),
                ('LINEABOVE',     (0, n-3), (-1, n-3),   1,   BORDER),
                ('PADDING',       (0, 0),   (-1, -1),    7),
                ('VALIGN',        (0, 0),   (-1, -1),    'MIDDLE'),
            ]))
            elements.append(KeepTogether(tbl))
            elements.append(Spacer(1, 7 * mm))

        # ── APPROVAL ─────────────────────────────────────────────────────
        approved_log = (invoice.audit_logs
                        .filter(action='approved')
                        .select_related('user')
                        .first())

        elements.append(Paragraph('APPROVAL', h2_s))
        approval = Table([[
            Paragraph('APPROVAL STATUS', lbl_s),
            Paragraph('APPROVED BY',     lbl_s),
            Paragraph('APPROVED ON',     lbl_s),
        ], [
            Paragraph(f'<font color="{status_hex}"><b>{status_lbl.upper()}</b></font>', bld_s),
            Paragraph(approved_log.user.full_name
                      if approved_log and approved_log.user else '—', val_s),
            Paragraph(approved_log.timestamp.strftime('%d %b %Y %H:%M')
                      if approved_log else '—', val_s),
        ]], colWidths=[usable_w/3, usable_w/3, usable_w/3])
        approval.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), BG_LIGHT),
            ('BOX',        (0, 0), (-1, -1), 0.5, BORDER),
            ('GRID',       (0, 0), (-1, -1), 0.5, BORDER),
            ('PADDING',    (0, 0), (-1, -1), 10),
            ('VALIGN',     (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        elements.append(approval)
        elements.append(Spacer(1, 7 * mm))

        # ── AUDIT TRAIL ───────────────────────────────────────────────────
        audit_logs = list(invoice.audit_logs.select_related('user').order_by('timestamp'))
        if audit_logs:
            elements.append(Paragraph('AUDIT TRAIL', h2_s))
            audit_rows = [[
                Paragraph('<b>Action</b>',       bld_s),
                Paragraph('<b>Performed By</b>', bld_s),
                Paragraph('<b>Date & Time</b>',  bld_s),
            ]]
            for log in audit_logs:
                audit_rows.append([
                    Paragraph(log.action.replace('_', ' ').title(), val_s),
                    Paragraph(log.user.full_name if log.user else 'System', val_s),
                    Paragraph(log.timestamp.strftime('%d %b %Y %H:%M'), val_s),
                ])
            audit_tbl = Table(audit_rows,
                              colWidths=[usable_w*0.32, usable_w*0.38, usable_w*0.30])
            audit_tbl.setStyle(TableStyle([
                ('BACKGROUND',    (0, 0), (-1, 0),  PRIMARY),
                ('TEXTCOLOR',     (0, 0), (-1, 0),  WHITE),
                ('ROWBACKGROUNDS',(0, 1), (-1, -1), [BG_LIGHT, WHITE]),
                ('GRID',          (0, 0), (-1, -1), 0.5, BORDER),
                ('PADDING',       (0, 0), (-1, -1), 8),
                ('VALIGN',        (0, 0), (-1, -1), 'MIDDLE'),
            ]))
            elements.append(KeepTogether(audit_tbl))

        # ── FOOTER — Generated By ─────────────────────────────────────────
        elements.append(Spacer(1, 8 * mm))
        elements.append(HRFlowable(width='100%', thickness=0.5, color=BORDER))
        elements.append(Spacer(1, 3 * mm))
        elements.append(Paragraph(
            f'Generated by InvoiceIQ &nbsp;|&nbsp; '
            f'Invoice ID: {invoice.id} &nbsp;|&nbsp; '
            f'Exported: {datetime.now().strftime("%d %b %Y %H:%M")}',
            sm_s
        ))

        doc.build(elements)
        buffer.seek(0)

        filename = f'invoice_{invoice.invoice_number or invoice.id}.pdf'
        response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response